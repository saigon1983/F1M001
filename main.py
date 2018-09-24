class Nation:
    import openpyxl
    from openpyxl.utils import get_column_letter

    WB = openpyxl.load_workbook(r"C:\Users\smeg\Desktop\F1M001\Data\Countries.xlsx")
    WS = WB["List1"]
    HEADERS = {}
    i = 1
    for header in WS[1]:
        HEADERS[header.value] = i
        i+=1

    COUNTRIES = {}
    i = 0
    for country in WS[get_column_letter(HEADERS['SHORT'])]:
        COUNTRIES[country.value] = i
        i+=1

    def __init__(self, nation):
        assert nation in Nation.COUNTRIES.keys(), "Sorry, wrong nation given"

        self.nameEng = Nation.WS[Nation.get_column_letter(Nation.HEADERS["NAME"])][Nation.COUNTRIES[nation]].value

        self.nameRus = []
        self.nameRus.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS"])][Nation.COUNTRIES[nation]].value)
        self.nameRus.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_2"])][Nation.COUNTRIES[nation]].value)
        self.nameRus.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_3"])][Nation.COUNTRIES[nation]].value)
        self.nameRus.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_4"])][Nation.COUNTRIES[nation]].value)
        self.nameRus.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_5"])][Nation.COUNTRIES[nation]].value)
        self.nameRus.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_6"])][Nation.COUNTRIES[nation]].value)

        self.nameEngAdj = Nation.WS[Nation.get_column_letter(Nation.HEADERS["ADJECTIVE"])][Nation.COUNTRIES[nation]].value
        self.nameRusAdj = []
        self.nameRusAdj.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_ADJECTIVE_M"])][Nation.COUNTRIES[nation]].value)
        self.nameRusAdj.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_ADJECTIVE_F"])][Nation.COUNTRIES[nation]].value)
        self.nameRusAdj.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_ADJECTIVE_N"])][Nation.COUNTRIES[nation]].value)
        self.nameRusAdj.append(Nation.WS[Nation.get_column_letter(Nation.HEADERS["RUS_ADJECTIVE_P"])][Nation.COUNTRIES[nation]].value)

    def nameEnglish(self):
        return self.nameEng

    def nameRussian(self, index = 0):
        return self.nameRus[index]

    def adjEnglish(self):
        return self.nameEngAdj

    def adjRussian(self, gender = 'M', capital = False):
        result = ""
        if gender == 'M':
            result = self.nameRusAdj[0]
        elif gender == 'F':
            result = self.nameRusAdj[1]
        elif gender == 'N':
            result = self.nameRusAdj[2]
        elif gender == 'P':
            result = self.nameRusAdj[3]
        else:
            raise ValueError('Wrong gender lexem passed!')
        if capital: result = result.capitalize()

        return  result

class Driver:
    def __init__(self, name, surname, gender, birthdate, nation):
        self.name = name
        self.surname = surname
        self.gender = gender
        self.birthdate = birthdate
        self.nation = nation

class Bolid:
    def __init__(self, teamName, motorName, sponsorName, nation):
        self.teamName = teamName
        self.motorName = motorName
        self.sponsorName = sponsorName
        self.nation = nation

class RaceUnit:
    def __init__(self, number, team, driver):
        self.number = number
        self.team = team
        self.driver = driver

nat1 = Nation('RUS')
print(nat1.nameEnglish())
print(nat1.nameRussian(3))
print(nat1.adjEnglish())
print(nat1.adjRussian('M', True))



